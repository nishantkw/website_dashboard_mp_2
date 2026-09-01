import json
import sys
from pathlib import Path
import openpyxl

def map_type(dt):
    if not dt:
        return 'text'
    d = str(dt).strip().lower().replace('\xa0', ' ')
    if 'timestamp' in d:
        return 'timestamptz'
    if d.startswith('bigint'):
        return 'bigint'
    if d.startswith('integer') or d == 'int':
        return 'integer'
    if d.startswith('smallint'):
        return 'smallint'
    if d.startswith('numeric') or 'decimal' in d:
        return 'numeric'
    if 'character varying' in d or d.startswith('varchar') or d == 'text':
        return 'text'
    if d.startswith('boolean'):
        return 'boolean'
    return 'text'

base = Path('Data dictionary') / 'Data dictionary'
tables = {}
for f in sorted(base.glob('*.xlsx')):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip().lower() if c else '' for c in rows[0]]

        def idx(*names):
            for n in names:
                for i, h in enumerate(header):
                    if n in h:
                        return i
            return None

        ti = idx('name of the table', 'table name')
        ci = idx('column_name', 'column name')
        di = idx('data type')
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            tname = str(row[ti]).strip() if ti is not None and row[ti] else None
            cname = str(row[ci]).strip() if ci is not None and row[ci] else None
            dtype = str(row[di]).strip() if di is not None and row[di] else 'text'
            if not tname or not cname or tname.lower() in ('none', '') or cname.lower() in ('none', ''):
                continue
            tables.setdefault(tname, [])
            if cname not in [c['name'] for c in tables[tname]]:
                tables[tname].append({'name': cname, 'type': map_type(dtype)})
    wb.close()

Path('backend').mkdir(exist_ok=True)
Path('schema_extract.json').write_text(json.dumps(tables, indent=2), encoding='utf-8')
print('Tables:', list(tables.keys()))
for t, cols in tables.items():
    print(f'  {t}: {len(cols)} cols')

SCHEMA_MAP = {
    't_bis_beneficiary_dtls': 'bis_raw',
    'claim_paid_excel_t': 'dmart_mp',
    'workflow_users_t': 'dmart_mp',
    't_workflow_transaction_audit': 'dmart_mp',
    'user_master_ump': 'ump_raw',
    'ump_user_dtl': 'ump_raw',
}

lines = [
    '-- Auto-generated from Data dictionary Excel files',
    '-- Compatible with PostgreSQL 14+ and Supabase',
    '',
    'CREATE SCHEMA IF NOT EXISTS bis_raw;',
    'CREATE SCHEMA IF NOT EXISTS dmart_mp;',
    'CREATE SCHEMA IF NOT EXISTS ump_raw;',
    'CREATE SCHEMA IF NOT EXISTS app_auth;',
    '',
]

for t, cols in tables.items():
    schema = SCHEMA_MAP.get(t, 'dmart_mp')
    lines.append(f'DROP TABLE IF EXISTS {schema}.{t} CASCADE;')
    lines.append(f'CREATE TABLE {schema}.{t} (')
    fixed = []
    used_pk = False
    for c in cols:
        name = c['name']
        typ = c['type']
        extra = ''
        if not used_pk and name in ('id_pk', 'work_id_pk'):
            extra = ' PRIMARY KEY'
            used_pk = True
        fixed.append(f'  {name} {typ}{extra}')
    if not used_pk:
        fixed.insert(0, '  id_pk bigserial PRIMARY KEY')
    lines.append(',\n'.join(fixed))
    lines.append(');')
    lines.append('')

lines += [
    'DROP TABLE IF EXISTS dmart_mp.t_suspicious_api_case_data CASCADE;',
    '''CREATE TABLE dmart_mp.t_suspicious_api_case_data (
  suspicion_id bigserial PRIMARY KEY,
  reference_number text,
  application_type text,
  entity_type text,
  entity_id text,
  hospital_name text,
  division_name text,
  district_name text,
  state_lgd_code text,
  fraud_type text,
  investigation_status text,
  amount_risk numeric,
  amount_recovered numeric,
  investigator text,
  lst_trigger_event_date timestamptz,
  crt_date timestamptz DEFAULT now(),
  updt_date timestamptz
);''',
    '',
    'DROP TABLE IF EXISTS dmart_mp.t_suspicious_api_case_dtls CASCADE;',
    '''CREATE TABLE dmart_mp.t_suspicious_api_case_dtls (
  id_pk bigserial PRIMARY KEY,
  state_lgd_code text,
  reference_number text,
  application_type text,
  vendor_id text,
  trigger_type text,
  trigger_code text,
  trigger_time timestamptz,
  trigger_reason text,
  crt_date timestamptz DEFAULT now(),
  crt_usr text,
  flag text,
  district_name text
);''',
    '',
    'DROP TABLE IF EXISTS dmart_mp.hospital_master CASCADE;',
    '''CREATE TABLE dmart_mp.hospital_master (
  id_pk bigserial PRIMARY KEY,
  hospital_code text,
  facility_id text,
  hospital_name text,
  hospital_type text,
  district_name text,
  division_name text,
  hosp_spec_type text,
  nabh_certified text,
  enrl_status text,
  active_status text,
  accreditation_status text,
  empaneled_date date,
  deempanel_status text,
  pgdnb_status text,
  bed_capacity integer
);''',
    '',
    'DROP TABLE IF EXISTS app_auth.dashboard_users CASCADE;',
    '''CREATE TABLE app_auth.dashboard_users (
  id serial PRIMARY KEY,
  username text UNIQUE NOT NULL,
  password_hash text NOT NULL,
  full_name text NOT NULL,
  role text NOT NULL,
  department text,
  active boolean DEFAULT true,
  created_at timestamptz DEFAULT now()
);''',
    '',
    'CREATE OR REPLACE VIEW dmart_mp.claim_paid_t AS SELECT * FROM dmart_mp.claim_paid_excel_t;',
    '',
    'CREATE INDEX IF NOT EXISTS idx_claim_district ON dmart_mp.claim_paid_excel_t (hosp_district_name);',
    'CREATE INDEX IF NOT EXISTS idx_claim_status ON dmart_mp.claim_paid_excel_t (case_status);',
    'CREATE INDEX IF NOT EXISTS idx_ben_dist ON bis_raw.t_bis_beneficiary_dtls (dist_name);',
    'CREATE INDEX IF NOT EXISTS idx_fraud_ref ON dmart_mp.t_suspicious_api_case_dtls (reference_number);',
    'CREATE INDEX IF NOT EXISTS idx_workflow_user ON dmart_mp.workflow_users_t (workflow_user);',
]

out = Path('backend/sql/001_init_schemas.sql')
out.parent.mkdir(parents=True, exist_ok=True)
out.write_text('\n'.join(lines), encoding='utf-8')
print('Wrote', out)
